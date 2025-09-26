# multi_tenant.py
# Contains views related to multi-tenant operations:
# - TenantViewSet
# - UploadSalaryDataAPIView
# - DownloadTemplateAPIView

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status, viewsets
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.http import HttpResponse
import pandas as pd
from datetime import datetime
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from ..models import (
    Tenant,
    SalaryData,
    EmployeeProfile,
)
from ..serializers import (
    TenantSerializer,
)
from ..utils.permissions import IsSuperUser
from ..utils.utils import (
    clean_decimal_value,
    clean_int_value,
    is_valid_name,
    validate_excel_columns,
    generate_employee_id,
)

TEMPLATE_COLUMNS = [
    "NAME",
    "SALARY",
    "ABSENT",
    "DAYS",
    "SL W/O OT",
    "OT",
    "HOUR RS",
    "CHARGES",
    "LATE",
    "CHARGE",
    "AMT",
    "SAL+OT",
    "25TH ADV",
    "OLD ADV",
    "NETT PAYABLE",
    "Department",
    "Total old ADV",
    "Balnce Adv",
    "INCENTIVE",
    "TDS",
    "SAL-TDS",
    "ADVANCE",
]

# Initialize logger
logger = logging.getLogger(__name__)

class TenantViewSet(viewsets.ModelViewSet):
    """
    Tenant management for super admins only
    """

    queryset = Tenant.objects.all()
    serializer_class = TenantSerializer
    permission_classes = [IsSuperUser]


class UploadSalaryDataAPIView(APIView):
    """
    Optimized salary data upload using exact template structure with bulk operations
    """

    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            # Get tenant from request
            tenant = getattr(request, "tenant", None)
            if not tenant:

                return Response(
                    {
                        "error": "No tenant found. Please ensure you're accessing the correct subdomain."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            excel_file = request.FILES.get("file")

            if not excel_file:

                return Response(
                    {"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
                )

            # Validate file extension

            if not excel_file.name.endswith((".xlsx", ".xls")):

                return Response(
                    {"error": "Please upload an Excel file (.xlsx or .xls)"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:

                # Read Excel file with NaN handling

                df = pd.read_excel(excel_file)

                # Replace all NaN values with appropriate defaults before processing

                df = df.fillna(
                    {
                        "NAME": "",
                        "SALARY": 0,
                        "ABSENT": 0,
                        "DAYS": 0,
                        "SL W/O OT": 0,
                        "OT": 0,
                        "HOUR RS": 0,
                        "CHARGES": 0,
                        "LATE": 0,
                        "CHARGE": 0,
                        "AMT": 0,
                        "SAL+OT": 0,
                        "25TH ADV": 0,
                        "OLD ADV": 0,
                        "NETT PAYABLE": 0,
                        "Department": "",
                        "Total old ADV": 0,
                        "Balnce Adv": 0,
                        "INCENTIVE": 0,
                        "TDS": 0,
                        "SAL-TDS": 0,
                        "ADVANCE": 0,
                    }
                )

                # Validate columns

                is_valid, error_message = validate_excel_columns(
                    df.columns.tolist(), TEMPLATE_COLUMNS
                )

                if not is_valid:

                    return Response(
                        {"error": error_message}, status=status.HTTP_400_BAD_REQUEST
                    )

                # Get month/year from request

                selected_month = request.data.get("month")

                selected_year = request.data.get("year")

                if not selected_month or not selected_year:

                    return Response(
                        {"error": "Please select month and year for the uploaded data"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Filter out invalid rows BEFORE processing

                valid_rows = df[df["NAME"].apply(is_valid_name)]

                if len(valid_rows) == 0:

                    return Response(
                        {"error": "No valid employee names found in the uploaded file"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Process data with bulk operations for maximum performance

                records_created = 0

                records_updated = 0

                errors = []

                # Collect all data for bulk operations

                salary_records_to_create = []

                salary_records_to_update = []

                employee_profiles_to_create = []

                # Get existing salary records for this period to avoid duplicates

                existing_salary_dict = {}

                existing_salaries = SalaryData.objects.filter(
                    tenant=tenant, year=int(selected_year), month=selected_month
                ).values("employee_id", "id")

                for salary in existing_salaries:

                    existing_salary_dict[salary["employee_id"]] = salary["id"]

                # Get existing employee profiles to avoid duplicates
                existing_employee_ids = set(
                    EmployeeProfile.objects.filter(tenant=tenant).values_list(
                        "employee_id", flat=True
                    )
                )

                # Get existing employees by name and department for reuse
                existing_employees_by_name = {}
                existing_employees = EmployeeProfile.objects.filter(tenant=tenant).values(
                    'employee_id', 'first_name', 'last_name', 'department'
                )
                for emp in existing_employees:
                    # Create a key based on name and department for matching
                    full_name = f"{emp['first_name']} {emp['last_name']}".strip()
                    key = f"{full_name}|{emp['department']}"
                    existing_employees_by_name[key] = emp['employee_id']

                # Prepare bulk data

                for index, row in valid_rows.iterrows():

                    try:

                        # Check if employee already exists by name and department
                        department = str(row.get("Department", "")).strip()
                        employee_name = str(row["NAME"]).strip()
                        
                        # Try to find existing employee by name and department
                        existing_employee_key = f"{employee_name}|{department}"
                        if existing_employee_key in existing_employees_by_name:
                            # Reuse existing employee ID
                            employee_id = existing_employees_by_name[existing_employee_key]
                        else:
                            # Generate new employee ID only if employee doesn't exist
                            employee_id = generate_employee_id(
                                row["NAME"], tenant.id, department
                            )

                        # Prepare salary data

                        salary_data = {
                            "tenant": tenant,
                            "employee_id": employee_id,
                            "year": int(selected_year),
                            "month": selected_month,
                            "name": str(row["NAME"]).strip(),
                            "salary": clean_decimal_value(row["SALARY"]),
                            "absent": clean_int_value(row["ABSENT"]),
                            "days": clean_int_value(row["DAYS"]),
                            "sl_wo_ot": clean_decimal_value(row["SL W/O OT"]),
                            "ot": clean_decimal_value(row["OT"]),
                            "hour_rs": clean_decimal_value(row["HOUR RS"]),
                            "charges": clean_decimal_value(row["CHARGES"]),
                            "late": clean_int_value(row["LATE"]),
                            "charge": clean_decimal_value(row["CHARGE"]),
                            "amt": clean_decimal_value(row["AMT"]),
                            "sal_ot": clean_decimal_value(row["SAL+OT"]),
                            "adv_25th": clean_decimal_value(row["25TH ADV"]),
                            "old_adv": clean_decimal_value(row["OLD ADV"]),
                            "nett_payable": clean_decimal_value(row["NETT PAYABLE"]),
                            "department": str(row.get("Department", "")).strip(),
                            "total_old_adv": clean_decimal_value(row["Total old ADV"]),
                            "balnce_adv": clean_decimal_value(row["Balnce Adv"]),
                            "incentive": clean_decimal_value(row["INCENTIVE"]),
                            "tds": clean_decimal_value(row["TDS"]),
                            "sal_tds": clean_decimal_value(row["SAL-TDS"]),
                            "advance": clean_decimal_value(row["ADVANCE"]),
                            "date": datetime(
                                int(selected_year),
                                self._get_month_number(selected_month),
                                1,
                            ).date(),
                        }

                        if employee_id in existing_salary_dict:

                            # Update existing record

                            salary_data["id"] = existing_salary_dict[employee_id]

                            salary_records_to_update.append(SalaryData(**salary_data))

                            records_updated += 1

                        else:

                            # Create new record

                            salary_records_to_create.append(SalaryData(**salary_data))

                            records_created += 1

                        # Prepare employee profile if doesn't exist
                        # Only create if this is a truly new employee (not found by name+department)
                        if (employee_id not in existing_employee_ids and 
                            existing_employee_key not in existing_employees_by_name):

                            existing_employee_ids.add(employee_id)

                            name_parts = str(row["NAME"]).strip().split()

                            employee_profiles_to_create.append(
                                EmployeeProfile(
                                    tenant=tenant,
                                    employee_id=employee_id,
                                    first_name=(
                                        name_parts[0] if name_parts else "Unknown"
                                    ),
                                    last_name=(
                                        " ".join(name_parts[1:])
                                        if len(name_parts) > 1
                                        else ""
                                    ),
                                    department=str(row.get("Department", "")).strip(),
                                    basic_salary=clean_decimal_value(row["SALARY"]),
                                    is_active=True,
                                )
                            )

                    except Exception as e:

                        errors.append(f"Row {index + 2}: {str(e)}")

                # Perform bulk operations

                with transaction.atomic():

                    # Deduplicate salary records to create (in case same employee appears multiple times in upload)
                    if salary_records_to_create:
                        # Create a dictionary to deduplicate by (employee_id, year, month)
                        unique_salary_records = {}
                        for record in salary_records_to_create:
                            key = (record.employee_id, record.year, record.month)
                            if key not in unique_salary_records:
                                unique_salary_records[key] = record
                            else:
                                # If duplicate found, keep the last one (most recent data)
                                unique_salary_records[key] = record
                        
                        # Convert back to list
                        salary_records_to_create = list(unique_salary_records.values())
                        records_created = len(salary_records_to_create)

                    # Bulk create new salary records

                    if salary_records_to_create:

                        SalaryData.objects.bulk_create(
                            salary_records_to_create, batch_size=100
                        )

                    # Bulk update existing salary records

                    if salary_records_to_update:

                        SalaryData.objects.bulk_update(
                            salary_records_to_update,
                            [
                                "name",
                                "salary",
                                "absent",
                                "days",
                                "sl_wo_ot",
                                "ot",
                                "hour_rs",
                                "charges",
                                "late",
                                "charge",
                                "amt",
                                "sal_ot",
                                "adv_25th",
                                "old_adv",
                                "nett_payable",
                                "department",
                                "total_old_adv",
                                "balnce_adv",
                                "incentive",
                                "tds",
                                "sal_tds",
                                "advance",
                                "date",
                            ],
                            batch_size=100,
                        )
                    # Deduplicate employee profiles to create (in case same employee appears multiple times in upload)
                    if employee_profiles_to_create:
                        # Create a dictionary to deduplicate by employee_id
                        unique_employee_profiles = {}
                        for profile in employee_profiles_to_create:
                            if profile.employee_id not in unique_employee_profiles:
                                unique_employee_profiles[profile.employee_id] = profile
                            else:
                                # If duplicate found, keep the last one (most recent data)
                                unique_employee_profiles[profile.employee_id] = profile
                        
                        # Convert back to list
                        employee_profiles_to_create = list(unique_employee_profiles.values())

                    # Bulk create new employee profiles
                    if employee_profiles_to_create:
                        EmployeeProfile.objects.bulk_create(
                            employee_profiles_to_create, batch_size=100
                        )

                return Response(
                    {
                        "message": "Upload completed successfully",
                        "records_created": records_created,
                        "records_updated": records_updated,
                        "total_processed": len(valid_rows),
                        "total_rows_in_file": len(df),
                        "errors": errors[:10],  # Show first 10 errors only
                        "total_errors": len(errors),
                    },
                    status=status.HTTP_200_OK,
                )

            except Exception as e:

                return Response(
                    {"error": f"Error processing Excel file: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        except Exception as e:

            return Response(
                {"error": f"Upload failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def _get_month_number(self, month_name):
        """Convert month name to number"""

        months = {
            "JANUARY": 1,
            "JAN": 1,
            "FEBRUARY": 2,
            "FEB": 2,
            "MARCH": 3,
            "MAR": 3,
            "APRIL": 4,
            "APR": 4,
            "MAY": 5,
            "JUNE": 6,
            "JUN": 6,
            "JULY": 7,
            "JUL": 7,
            "AUGUST": 8,
            "AUG": 8,
            "SEPTEMBER": 9,
            "SEP": 9,
            "OCTOBER": 10,
            "OCT": 10,
            "NOVEMBER": 11,
            "NOV": 11,
            "DECEMBER": 12,
            "DEC": 12,
        }

        return months.get(month_name.upper(), 1)


class DownloadTemplateAPIView(APIView):
    """

    Download the Excel template for salary data upload

    """

    permission_classes = [IsAuthenticated]

    def get(self, request):

        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Data Template"
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers

        for col_idx, column in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Add sample data

        sample_data = [
            [
                "John Doe",
                50000,
                2,
                28,
                45000,
                5,
                208.33,
                1041.65,
                30,
                6.94,
                208.33,
                46041.65,
                5000,
                2000,
                39041.65,
                "IT",
                10000,
                8000,
                1000,
                500,
                38541.65,
                0,
            ],
            [
                "Jane Smith",
                45000,
                1,
                29,
                43500,
                3,
                187.5,
                562.5,
                15,
                6.25,
                93.75,
                44062.5,
                4000,
                1500,
                38562.5,
                "HR",
                8000,
                6500,
                800,
                400,
                38162.5,
                0,
            ],
        ]

        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx > 1:  # Numeric columns
                    cell.alignment = Alignment(horizontal="right")

        # Adjust column widths

        for col_idx, column in enumerate(TEMPLATE_COLUMNS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
                len(column) + 2, 12
            )

        # Create response

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="salary_data_template.xlsx"'
        )
        wb.save(response)
        return response
